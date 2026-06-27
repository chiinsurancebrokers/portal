"""
xlsx_to_csv.py
───────────────────────────────────────────────────────────────────────────
Converts an .xlsx export (old portal / insurance system) into a .csv file
that the CHI portal's "Import Ληξιαρίου" page (/agent/import) will parse
correctly.

WHY THIS EXISTS
The import page only reads .csv/.txt files, but exports from the old
system / Excel sometimes arrive as .xlsx. This script reads the first
sheet of an .xlsx file, figures out which of the portal's 3 known column
layouts it matches, and writes a .csv using exactly the delimiter, header
text, and date/number formatting that the portal's parser
(_parse_lixiario_csv in app.py) expects:

    Format A — Ληξιάρια   (old "ληξιάριο" export, ; delimited, title+header row)
        Έναρξη;Λήξη;Εταιρεία;Κλάδος;Συμβόλαιο;Απόδειξη;Είδος;Χαρακτ/κό;
        Πελάτης;ΑΦΜ;Τηλέφωνο;Κινητό;Συνεργάτης;Ημ.Εκτύπ.;Μικτά;Καθαρά

    Format B — Παραγωγή   (production export, ; delimited, title+header row)
        Χαρακτ/κό;Πελάτης;Συμβόλαιο;Απόδειξη;Κατηγορία;Κλάδος;Εταιρεία;
        Έκδοση;Έναρξη;Λήξη;Μικτά;Καθαρά;Εξ.Προμήθεια;Υπόλ.

    Format C — Ανανεώσεις (renewals export, , delimited, single header row)
        Συμβόλαιο,Χαρακτηριστικό,Πελάτης,Κλάδος,Εταιρεία,Έναρξη,Λήξη,
        Μικτά,Τηλέφωνο,Κινητό

Detection is based on matching the xlsx's own header row (case-insensitive,
order-independent) against each known layout's column names. If your xlsx
already has one of these exact header rows (any column order), the script
maps your columns into the right positions automatically. If it doesn't
recognize the header, it tells you instead of guessing — a silently wrong
guess is worse than no conversion at all.

Dates are normalized to DD/MM/YYYY text and amounts to "1234,56" Greek
decimal-comma style (matching the legacy export convention the parser is
built around), regardless of how Excel stored them.

USAGE
    python3 xlsx_to_csv.py INPUT.xlsx [OUTPUT.csv] [--sheet NAME]

    If OUTPUT.csv is omitted, it's written next to the input file with a
    .csv extension. If the workbook has more than one sheet, the first
    sheet is used unless --sheet is given.

REQUIREMENTS
    pip install openpyxl
"""
import sys
import os
import argparse
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("This script needs openpyxl. Install it with:\n    pip install openpyxl --break-system-packages")
    sys.exit(1)


# ── Known column layouts ────────────────────────────────────────────────
# Each entry maps the layout's expected header names (in the app's exact
# column order) to: delimiter, whether the app expects a title row above
# the header, and which column indices are dates / amounts (so we can
# reformat them no matter how Excel stored them).

FORMAT_A = {
    "name": "lixiario",
    "headers": ["Έναρξη", "Λήξη", "Εταιρεία", "Κλάδος", "Συμβόλαιο",
                "Απόδειξη", "Είδος", "Χαρακτ/κό", "Πελάτης", "ΑΦΜ",
                "Τηλέφωνο", "Κινητό", "Συνεργάτης", "Ημ.Εκτύπ.",
                "Μικτά", "Καθαρά"],
    "delim": ";",
    "title_row": "Ληξιάρια",   # the app skips 2 rows for this format
    "date_cols": [0, 1, 13],
    "amount_cols": [14, 15],
}

FORMAT_B = {
    "name": "paragogi",
    "headers": ["Χαρακτ/κό", "Πελάτης", "Συμβόλαιο", "Απόδειξη",
                "Κατηγορία", "Κλάδος", "Εταιρεία", "Έκδοση",
                "Έναρξη", "Λήξη", "Μικτά", "Καθαρά",
                "Εξ.Προμήθεια", "Υπόλ."],
    "delim": ";",
    "title_row": "Παραγωγή",   # the app skips 2 rows for this format
    "date_cols": [7, 8, 9],
    "amount_cols": [10, 11, 12, 13],
}

FORMAT_C = {
    "name": "ananeoseis",
    "headers": ["Συμβόλαιο", "Χαρακτηριστικό", "Πελάτης", "Κλάδος",
                "Εταιρεία", "Έναρξη", "Λήξη", "Μικτά", "Τηλέφωνο",
                "Κινητό"],
    "delim": ",",
    "title_row": None,         # the app expects only 1 header row
    "date_cols": [5, 6],
    "amount_cols": [7],
}

KNOWN_FORMATS = [FORMAT_A, FORMAT_B, FORMAT_C]


def _norm(s):
    """Normalize a header cell for comparison: casefold, strip accents-insensitive enough,
    collapse whitespace."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = s.replace("Ά", "Α").replace("Έ", "Ε").replace("Ή", "Η").replace("Ί", "Ι") \
         .replace("Ό", "Ο").replace("Ύ", "Υ").replace("Ώ", "Ω")
    s = " ".join(s.split())
    return s


def detect_format(header_cells):
    """Compare the xlsx header row against each known layout (order-independent).
    Returns (format_dict, column_index_map) where column_index_map[i] is the
    source column index in the xlsx for the i-th column of the matched layout,
    or (None, None) if nothing matches well enough."""
    norm_header = [_norm(c) for c in header_cells]

    best_fmt, best_map, best_score = None, None, 0
    for fmt in KNOWN_FORMATS:
        wanted = [_norm(h) for h in fmt["headers"]]
        col_map = {}
        matched = 0
        for i, want in enumerate(wanted):
            if want in norm_header:
                col_map[i] = norm_header.index(want)
                matched += 1
        # require at least 70% of the layout's columns present to call it a match
        if matched >= max(3, int(0.7 * len(wanted))) and matched > best_score:
            best_fmt, best_map, best_score = fmt, col_map, matched

    return best_fmt, best_map


def _to_date_str(value):
    """Normalize a cell to DD/MM/YYYY text, however Excel stored it."""
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    s = str(value).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s.split(" ")[0], fmt).strftime("%d/%m/%Y")
        except Exception:
            pass
    return s  # leave as-is if unrecognized; better than dropping data


def _to_amount_str(value, delim=";"):
    """Normalize a cell to a numeric string, stripping currency symbols.
    Uses comma-decimal ('1234,56') for ';'-delimited formats, matching the
    legacy export convention — but dot-decimal ('1234.56') for ','-delimited
    formats, since a literal comma inside a field would corrupt a naive
    comma-split (the app's parser does not do quoted-CSV parsing)."""
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return f"{value:.2f}" if delim == "," else f"{value:.2f}".replace(".", ",")
    s = str(value).strip()
    s = s.replace("€", "").replace("EUR", "").replace("ευρώ", "").strip()
    if not s:
        return ""
    # normalize whatever decimal style it's already in to a plain float first
    try:
        if "," in s and "." in s:
            # e.g. "1.234,56" -> European thousands+decimal
            num = float(s.replace(".", "").replace(",", "."))
        elif "," in s:
            num = float(s.replace(",", "."))
        else:
            num = float(s)
        return f"{num:.2f}" if delim == "," else f"{num:.2f}".replace(".", ",")
    except Exception:
        return s


def convert(input_path, output_path=None, sheet_name=None):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        print("The sheet is empty — nothing to convert.")
        sys.exit(1)

    header_cells = [c for c in all_rows[0]]
    fmt, col_map = detect_format(header_cells)

    if not fmt:
        print("Could not recognize the column layout of this file.")
        print(f"Header row found: {list(header_cells)}")
        print("\nThis script only knows the 3 layouts the portal's import page")
        print("supports (Ληξιάρια, Παραγωγή, Ανανεώσεις). If this is a new")
        print("export format, the import page's parser needs a new format")
        print("branch added for it too — converting to CSV alone won't help")
        print("until the app knows how to read these columns.")
        sys.exit(1)

    print(f"Detected format: {fmt['name']}  (delimiter '{fmt['delim']}', "
          f"{'title+header row' if fmt['title_row'] else 'single header row'})")

    data_rows = all_rows[1:]

    out_rows = []
    collisions = 0
    for row in data_rows:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        out = []
        for i, header_name in enumerate(fmt["headers"]):
            src_idx = col_map.get(i)
            val = row[src_idx] if (src_idx is not None and src_idx < len(row)) else None
            if i in fmt["date_cols"]:
                cell = _to_date_str(val)
            elif i in fmt["amount_cols"]:
                cell = _to_amount_str(val, fmt["delim"])
            else:
                cell = "" if val is None else str(val).strip()
            # The app's parser does a plain str.split(delim) — it does not
            # understand quoted CSV fields. A stray delimiter character
            # inside a text field (e.g. a company name like "Α.Ε., Αθήνα"
            # in a comma-delimited file) would silently shift every column
            # after it. Replace it rather than let that happen quietly.
            if fmt["delim"] in cell:
                cell = cell.replace(fmt["delim"], " ")
                collisions += 1
            out.append(cell)
        out_rows.append(out)

    if collisions:
        print(f"⚠ Replaced the delimiter character inside {collisions} field(s) "
              f"that contained it (e.g. a comma in a company name), to avoid "
              f"shifting columns — please review the output.")

    if output_path is None:
        base, _ = os.path.splitext(input_path)
        output_path = base + ".csv"

    delim = fmt["delim"]
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        if fmt["title_row"]:
            f.write(fmt["title_row"] + delim * (len(fmt["headers"]) - 1) + "\n")
        f.write(delim.join(fmt["headers"]) + "\n")
        for row in out_rows:
            f.write(delim.join(row) + "\n")

    print(f"Wrote {len(out_rows)} data row(s) to: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Convert an .xlsx renewals/production export to a .csv the portal's import page can read."
    )
    parser.add_argument("input", help="Path to the .xlsx file")
    parser.add_argument("output", nargs="?", default=None, help="Path to write the .csv (optional)")
    parser.add_argument("--sheet", default=None, help="Sheet name to read (default: first sheet)")
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        print(f"File not found: {args.input}")
        sys.exit(1)

    convert(args.input, args.output, args.sheet)


if __name__ == "__main__":
    main()
